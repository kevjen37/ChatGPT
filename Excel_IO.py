import os
import openai
import requests
import openpyxl

#***************************************************************
#USER DEFINED INPUTS:

api_key = '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' #fake key
path = '/Users/kevinjennings/Documents/ChatGPT_API/Test.xlsx'

#read cell:
r_row = 1
r_col = 1

#write cell:
w_row = 2
w_col = 1

#***************************************************************


openai.api_key = api_key

wb_obj = openpyxl.load_workbook(path)

wb_obj.save(path)
wb_obj.close()

sheet_obj = wb_obj.active
cell_obj = sheet_obj.cell(row = r_row, column = r_col).value

text_input = str(cell_obj)
print('\n', text_input, '\n')

completion = openai.ChatCompletion.create(
    model="gpt-3.5-turbo",
    messages = [{"role": "user", "content": text_input}],
    max_tokens = 1024,
    temperature = 0.8)

message = completion.choices[0].message.content

cell_output = sheet_obj.cell(row = w_row, column = w_col)
cell_output.value = message

wb_obj.save(path)


print(message)
